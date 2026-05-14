"""삼성증권 거래내역서 xlsx 파서."""

from __future__ import annotations

import io
import re

import openpyxl

from invest_note_api.domain.trade_types import TRADE_TYPE_BUY, TRADE_TYPE_SELL

from .base import BrokerStatementParser, ParsedTrade, ParseResult, parse_number

_BUY_NAMES = {"매수", "매수_NXT"}
_SELL_NAMES = {"매도", "매도_NXT"}
_TRADE_NAMES = _BUY_NAMES | _SELL_NAMES

_ACCOUNT_RE = re.compile(r"(\d{7,}-\d{2,})")
_SHEET_NAME = "Col1"
_HEADER_ROW = 2   # 1-indexed
_DATA_START = 3


class SamsungXlsxParser(BrokerStatementParser):
    key = "samsung_xlsx"
    display_name = "삼성증권"

    @classmethod
    def match(cls, filename: str, head_bytes: bytes) -> bool:
        if re.match(r"^삼성증권.*\.xlsx?$", filename, re.IGNORECASE):
            return True
        # xlsx 매직 바이트(PK zip) + 시트명 시그니처로 fallback
        if head_bytes[:2] == b"PK" and b"Col1" in head_bytes[:4096]:
            return True
        return False

    def parse(self, file_bytes: bytes, filename: str) -> ParseResult:
        result = ParseResult()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

        sheet = wb[_SHEET_NAME] if _SHEET_NAME in wb.sheetnames else wb.active

        # A1 셀에서 계좌번호 추출
        rows_iter = sheet.iter_rows(values_only=True)
        first_row = next(rows_iter, ())
        if first_row:
            a1 = str(first_row[0] or "")
            m = _ACCOUNT_RE.search(a1)
            if m:
                result.account_hint = m.group(1)

        # 헤더 행 (2행)
        header_row = next(rows_iter, ())
        headers = [str(c).strip() if c is not None else "" for c in header_row]

        def col(row: tuple, name: str):
            try:
                idx = headers.index(name)
                return row[idx] if idx < len(row) else None
            except ValueError:
                return None

        for sheet_row_no, row in enumerate(rows_iter, start=_DATA_START):
            trade_name = str(col(row, "거래명") or "").strip()
            if not trade_name:
                continue
            if trade_name not in _TRADE_NAMES:
                if trade_name:  # 배당금입금 등 비거래 행
                    result.add_error(sheet_row_no, f"미지원 거래명: {trade_name}", {"거래명": trade_name})
                continue

            trade_type = TRADE_TYPE_BUY if trade_name in _BUY_NAMES else TRADE_TYPE_SELL

            asset_name = str(col(row, "종목명") or "").strip()
            if not asset_name:
                result.add_error(sheet_row_no, "종목명 없음")
                continue

            traded_at_raw = col(row, "거래일자")
            if traded_at_raw is None:
                result.add_error(sheet_row_no, "거래일자 없음")
                continue
            traded_at_kst = str(traded_at_raw).strip()[:10]  # "YYYY-MM-DD"

            quantity = parse_number(col(row, "거래수량") or 0)
            price = parse_number(col(row, "거래단가") or 0)
            commission = parse_number(col(row, "수수료/Fee") or 0)
            tax = parse_number(col(row, "제세금/대출이자") or 0)

            if quantity <= 0 or price <= 0:
                result.add_error(sheet_row_no, f"수량({quantity}) 또는 단가({price})가 0 이하")
                continue

            currency_raw = str(col(row, "통화코드") or "").strip()
            currency = currency_raw if currency_raw in ("KRW", "USD") else "KRW"
            if currency == "USD":
                result.usd_skip_count += 1
                result.add_error(sheet_row_no, "USD 거래 — MVP 미지원", {"거래명": trade_name, "종목명": asset_name})
                continue

            result.trades.append(
                ParsedTrade(
                    source_row_no=sheet_row_no,
                    traded_at_kst=traded_at_kst,
                    trade_type=trade_type,
                    asset_name=asset_name,
                    quantity=quantity,
                    price=price,
                    commission=commission,
                    tax=tax,
                    currency=currency,
                    account_hint=result.account_hint,
                    raw={"거래명": trade_name, "종목명": asset_name},
                )
            )

        wb.close()
        return result
