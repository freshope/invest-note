"""증권사 파서 단위 테스트 (synthetic in-memory 파일 사용)."""

import io

import openpyxl
import pytest

from invest_note_api.broker_import import PARSERS, detect_broker
from invest_note_api.broker_import.samsung_xlsx import SamsungXlsxParser
from invest_note_api.broker_import.toss_pdf import TossPdfParser, _parse_ticker_hint
from invest_note_api.broker_import.base import ParseResult


def _make_samsung_xlsx(rows: list[dict], account_meta: str = "7157197877-14 [ ISA ] 홍길동") -> bytes:
    """테스트용 삼성증권 xlsx 파일을 메모리에서 생성한다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Col1"

    # A1: 계좌 메타
    ws["A1"] = account_meta
    ws["B1"] = "< 2026.01.01 ~ 2026.12.31 >"

    # 2행: 헤더
    headers = [
        "거래일자", "거래명", "거래수량", "거래금액", "제세금/대출이자",
        "현금잔액", "상대계좌명", "변제금액", "통화코드", "외화정산금액",
        "거래번호", "종목명", "거래단가", "정산금액", "수수료/Fee",
        "잔고수량/펀드평가금액", "상대계좌번호", "신용/대출금",
        "외화거래금액", "외화예수금액", "처리점",
    ]
    ws.append(headers)

    # 3행~: 데이터
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _buy_row(date_str="2026-03-30", name="매수", asset="삼성전자",
             qty=10, price=70000, fee=22, tax=0, currency="") -> dict:
    return {
        "거래일자": date_str,
        "거래명": name,
        "종목명": asset,
        "거래수량": qty,
        "거래단가": price,
        "거래금액": qty * price,
        "수수료/Fee": fee,
        "제세금/대출이자": tax,
        "통화코드": currency,
    }


class TestSamsungXlsxParser:
    parser = SamsungXlsxParser()

    def test_match_by_filename(self):
        assert SamsungXlsxParser.match("삼성증권 거래내역서.xlsx", b"PK")
        assert not SamsungXlsxParser.match("토스증권_거래내역.pdf", b"%PDF")

    def test_parses_buy_trade(self):
        xlsx = _make_samsung_xlsx([_buy_row()])
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert len(result.trades) == 1
        t = result.trades[0]
        assert t.trade_type == "BUY"
        assert t.asset_name == "삼성전자"
        assert t.quantity == 10
        assert t.price == 70000
        assert t.commission == 22
        assert t.currency == "KRW"

    def test_parses_sell_trade(self):
        xlsx = _make_samsung_xlsx([_buy_row(name="매도", tax=350)])
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert len(result.trades) == 1
        assert result.trades[0].trade_type == "SELL"
        assert result.trades[0].tax == 350

    def test_nxt_trades_are_included(self):
        xlsx = _make_samsung_xlsx([
            _buy_row(name="매수_NXT"),
            _buy_row(name="매도_NXT"),
        ])
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert len(result.trades) == 2

    def test_non_trade_rows_are_errors(self):
        xlsx = _make_samsung_xlsx([
            _buy_row(name="배당금입금"),
            _buy_row(name="이체입금"),
            _buy_row(),
        ])
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert len(result.trades) == 1
        assert len(result.errors) == 2

    def test_usd_trade_skipped(self):
        row = _buy_row(currency="USD")
        xlsx = _make_samsung_xlsx([row])
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert len(result.trades) == 0
        assert result.usd_skip_count == 1

    def test_account_hint_extracted(self):
        xlsx = _make_samsung_xlsx([], account_meta="7157197877-14 [ ISA ] 홍길동")
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert result.account_hint == "7157197877-14"

    def test_zero_quantity_produces_error(self):
        row = _buy_row(qty=0)
        xlsx = _make_samsung_xlsx([row])
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert len(result.trades) == 0
        assert len(result.errors) == 1

    def test_multiple_trades(self):
        xlsx = _make_samsung_xlsx([
            _buy_row(date_str="2026-01-10", name="매수", asset="삼성전자", qty=5, price=70000),
            _buy_row(date_str="2026-02-01", name="매도", asset="삼성전자", qty=3, price=72000),
            _buy_row(date_str="2026-03-15", name="매수", asset="SK하이닉스", qty=10, price=150000),
        ])
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert len(result.trades) == 3
        assert result.trades[0].asset_name == "삼성전자"
        assert result.trades[1].trade_type == "SELL"
        assert result.trades[2].asset_name == "SK하이닉스"


class TestParseTickerHint:
    def test_standard_ticker(self):
        name, code = _parse_ticker_hint("삼성전자(005930)")
        assert name == "삼성전자"
        assert code == "005930"

    def test_preferred_share_a_prefix(self):
        name, code = _parse_ticker_hint("삼성전자우(A005935)")
        assert name == "삼성전자우"
        assert code == "005935"

    def test_no_code_returns_none(self):
        name, code = _parse_ticker_hint("삼성전자")
        assert name == "삼성전자"
        assert code is None

    def test_strips_whitespace(self):
        name, code = _parse_ticker_hint("  SK하이닉스(000660)  ")
        assert name == "SK하이닉스"
        assert code == "000660"


def _toss_buy_line(
    date_str="2026-03-30",
    name="삼성전자(A005930)",
    qty=10,
    amount=700000,
    price=70000,
    fee=22,
    tax=0,
    sec_tax=1260,
) -> str:
    """BUY 행 텍스트 (page.extract_text() 한 줄 포맷)."""
    return f"{date_str} 구매 {name} {qty} {amount} {price} {fee} {tax} {sec_tax} 0 0 0"


def _toss_sell_line(
    date_str="2026-04-15",
    name="삼성전자(A005930)",
    qty=5,
    amount=365000,
    price=73000,
    fee=51,
    tax=730,
    sec_tax=365,
) -> str:
    """SELL 행 텍스트."""
    return f"{date_str} 판매 {name} {qty} {amount} {price} {fee} {tax} {sec_tax} 0 0 0"


class TestTossPdfParserLine:
    parser = TossPdfParser()

    def test_parses_buy_line(self):
        result = ParseResult()
        trade = self.parser._parse_line(_toss_buy_line(), 1, result)
        assert trade is not None
        assert trade.trade_type == "BUY"
        assert trade.asset_name == "삼성전자"
        assert trade.ticker_hint == "005930"
        assert trade.quantity == 10
        assert trade.price == pytest.approx(70000.0)
        assert trade.commission == pytest.approx(22.0)
        assert trade.tax == pytest.approx(1260.0)

    def test_parses_sell_line(self):
        result = ParseResult()
        trade = self.parser._parse_line(_toss_sell_line(), 2, result)
        assert trade is not None
        assert trade.trade_type == "SELL"
        assert trade.asset_name == "삼성전자"
        assert trade.quantity == 5
        assert trade.price == pytest.approx(73000.0)
        assert trade.commission == pytest.approx(51.0)
        assert trade.tax == pytest.approx(1095.0)

    def test_zero_quantity_produces_error(self):
        result = ParseResult()
        trade = self.parser._parse_line(_toss_buy_line(qty=0), 4, result)
        assert trade is None
        assert any("수량" in e["reason"] for e in result.errors)

    def test_dot_date_separator(self):
        result = ParseResult()
        trade = self.parser._parse_line(_toss_buy_line(date_str="2026.03.30"), 5, result)
        assert trade is not None
        assert trade.traded_at_kst == "2026-03-30"

    def test_price_fallback_from_amount_when_zero(self):
        result = ParseResult()
        line = _toss_buy_line(qty=4, amount=400000, price=0)
        trade = self.parser._parse_line(line, 7, result)
        assert trade is not None
        assert trade.price == pytest.approx(100000.0)

    def test_match_by_filename(self):
        assert TossPdfParser.match("토스증권_거래내역서_20250417_20260416_1.pdf", b"%PDF")
        assert not TossPdfParser.match("삼성증권 거래내역서.xlsx", b"PK")


class TestTossPdfParserFixture:
    """실제 sample PDF 로 회귀 가드. extract_tables 가 비어 있을 때 0건 침묵 반환되던
    버그를 다시 잡기 위한 통합 테스트."""

    parser = TossPdfParser()

    @pytest.fixture
    def sample_pdf_bytes(self) -> bytes:
        from pathlib import Path
        path = Path(__file__).resolve().parents[2] / "sample" / "토스증권_거래내역서_20250417_20260416_1.pdf"
        if not path.exists():
            pytest.skip(f"sample PDF not present: {path}")
        return path.read_bytes()

    def test_parses_real_sample(self, sample_pdf_bytes: bytes):
        result = self.parser.parse(sample_pdf_bytes, "토스증권_거래내역서_20250417_20260416_1.pdf")
        assert len(result.trades) > 0
        assert result.account_hint == "101-01-024891"
        # BUY/SELL 둘 다 추출되어야 한다.
        types = {t.trade_type for t in result.trades}
        assert "BUY" in types
        assert "SELL" in types
        # 모든 거래에서 단가/수량 > 0.
        for t in result.trades:
            assert t.price > 0
            assert t.quantity > 0


class TestDetectBroker:
    def test_detects_samsung_by_filename(self):
        xlsx_bytes = _make_samsung_xlsx([])
        key = detect_broker("삼성증권 거래내역서.xlsx", xlsx_bytes)
        assert key == "samsung_xlsx"

    def test_detects_toss_by_filename(self):
        # 실제 PDF 바이트 없이 파일명으로만 감지
        fake_pdf = b"%PDF-1.4 " + b"\x00" * 100
        key = detect_broker("토스증권_거래내역서_20250417_20260416_1.pdf", fake_pdf)
        assert key == "toss_pdf"

    def test_returns_none_for_unknown(self):
        key = detect_broker("unknown_bank.xlsx", b"PK\x03\x04")
        assert key is None

    def test_parsers_registry_has_both_brokers(self):
        assert "samsung_xlsx" in PARSERS
        assert "toss_pdf" in PARSERS
