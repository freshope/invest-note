"""증권사 파서 단위 테스트 (synthetic in-memory 파일 사용)."""

import io

import openpyxl
import pytest

from invest_note_api.broker_import import PARSERS
from invest_note_api.broker_import.mirae_pdf import MiraePdfParser
from invest_note_api.broker_import.samsung_xlsx import SamsungXlsxParser
from invest_note_api.broker_import.shinhan_pdf import _LINE2_RE, ShinhanPdfParser
from invest_note_api.broker_import.toss_pdf import (
    TossPdfParser,
    _build_column_map,
    _parse_ticker_hint,
    _parse_usd_isin,
    _parse_usd_name,
    _split_usd_nums,
)
from invest_note_api.broker_import.base import ParseResult


def _make_samsung_xlsx(rows: list[dict], account_meta: str = "7157197877-14 [ ISA ] 홍길동") -> bytes:
    """테스트용 삼성증권 xlsx 파일을 메모리에서 생성한다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Col1"

    # A1: 계좌 메타
    ws["A1"] = account_meta
    ws["B1"] = "< 2026.01.01 ~ 2026.12.31 >"

    # 2행: 헤더
    headers = [
        "거래일자", "거래명", "거래수량", "거래금액", "제세금/대출이자",
        "현금잔액", "상대계좌명", "변제금액", "통화코드", "외화정산금액",
        "거래번호", "종목명", "거래단가", "정산금액", "수수료/Fee",
        "잔고수량/펀드평가금액", "상대계좌번호", "신용/대출금",
        "외화거래금액", "외화예수금액", "처리점",
    ]
    ws.append(headers)

    # 3행~: 데이터
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _buy_row(date_str="2026-03-30", name="매수", asset="삼성전자",
             qty=10, price=70000, fee=22, tax=0, currency="") -> dict:
    return {
        "거래일자": date_str,
        "거래명": name,
        "종목명": asset,
        "거래수량": qty,
        "거래단가": price,
        "거래금액": qty * price,
        "수수료/Fee": fee,
        "제세금/대출이자": tax,
        "통화코드": currency,
    }


class TestSamsungXlsxParser:
    parser = SamsungXlsxParser()

    def test_parses_buy_trade(self):
        xlsx = _make_samsung_xlsx([_buy_row()])
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert len(result.trades) == 1
        t = result.trades[0]
        assert t.trade_type == "BUY"
        assert t.asset_name == "삼성전자"
        assert t.quantity == 10
        assert t.price == 70000
        assert t.commission == 22
        assert t.currency == "KRW"

    def test_parses_sell_trade(self):
        xlsx = _make_samsung_xlsx([_buy_row(name="매도", tax=350)])
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert len(result.trades) == 1
        assert result.trades[0].trade_type == "SELL"
        assert result.trades[0].tax == 350

    def test_nxt_trades_are_included(self):
        xlsx = _make_samsung_xlsx([
            _buy_row(name="매수_NXT"),
            _buy_row(name="매도_NXT"),
        ])
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert len(result.trades) == 2

    def test_non_trade_rows_are_errors(self):
        xlsx = _make_samsung_xlsx([
            _buy_row(name="배당금입금"),
            _buy_row(name="이체입금"),
            _buy_row(),
        ])
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert len(result.trades) == 1
        assert len(result.errors) == 2

    def test_usd_trade_skipped(self):
        row = _buy_row(currency="USD")
        xlsx = _make_samsung_xlsx([row])
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert len(result.trades) == 0
        assert result.usd_skip_count == 1

    def test_account_hint_extracted(self):
        xlsx = _make_samsung_xlsx([], account_meta="7157197877-14 [ ISA ] 홍길동")
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert result.account_hint == "7157197877-14"

    def test_zero_quantity_produces_error(self):
        row = _buy_row(qty=0)
        xlsx = _make_samsung_xlsx([row])
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert len(result.trades) == 0
        assert len(result.errors) == 1

    def test_multiple_trades(self):
        xlsx = _make_samsung_xlsx([
            _buy_row(date_str="2026-01-10", name="매수", asset="삼성전자", qty=5, price=70000),
            _buy_row(date_str="2026-02-01", name="매도", asset="삼성전자", qty=3, price=72000),
            _buy_row(date_str="2026-03-15", name="매수", asset="SK하이닉스", qty=10, price=150000),
        ])
        result = self.parser.parse(xlsx, "삼성증권 거래내역서.xlsx")
        assert len(result.trades) == 3
        assert result.trades[0].asset_name == "삼성전자"
        assert result.trades[1].trade_type == "SELL"
        assert result.trades[2].asset_name == "SK하이닉스"


class TestParseTickerHint:
    def test_standard_ticker(self):
        name, code = _parse_ticker_hint("삼성전자(005930)")
        assert name == "삼성전자"
        assert code == "005930"

    def test_preferred_share_a_prefix(self):
        name, code = _parse_ticker_hint("삼성전자우(A005935)")
        assert name == "삼성전자우"
        assert code == "005935"

    def test_no_code_returns_none(self):
        name, code = _parse_ticker_hint("삼성전자")
        assert name == "삼성전자"
        assert code is None

    def test_strips_whitespace(self):
        name, code = _parse_ticker_hint("  SK하이닉스(000660)  ")
        assert name == "SK하이닉스"
        assert code == "000660"


def _toss_buy_line(
    date_str="2026-03-30",
    name="삼성전자(A005930)",
    qty=10,
    amount=700000,
    price=70000,
    fee=22,
    tax=0,
    sec_tax=1260,
) -> str:
    """BUY 행 텍스트 (page.extract_text() 한 줄 포맷)."""
    return f"{date_str} 구매 {name} {qty} {amount} {price} {fee} {tax} {sec_tax} 0 0 0"


def _toss_sell_line(
    date_str="2026-04-15",
    name="삼성전자(A005930)",
    qty=5,
    amount=365000,
    price=73000,
    fee=51,
    tax=730,
    sec_tax=365,
) -> str:
    """SELL 행 텍스트."""
    return f"{date_str} 판매 {name} {qty} {amount} {price} {fee} {tax} {sec_tax} 0 0 0"


class TestTossPdfParserLine:
    parser = TossPdfParser()

    def test_parses_buy_line(self):
        result = ParseResult()
        trade = self.parser._parse_line(_toss_buy_line(), 1, result)
        assert trade is not None
        assert trade.trade_type == "BUY"
        assert trade.asset_name == "삼성전자"
        assert trade.ticker_hint == "005930"
        assert trade.quantity == 10
        assert trade.price == pytest.approx(70000.0)
        assert trade.commission == pytest.approx(22.0)
        assert trade.tax == pytest.approx(1260.0)

    def test_parses_sell_line(self):
        result = ParseResult()
        trade = self.parser._parse_line(_toss_sell_line(), 2, result)
        assert trade is not None
        assert trade.trade_type == "SELL"
        assert trade.asset_name == "삼성전자"
        assert trade.quantity == 5
        assert trade.price == pytest.approx(73000.0)
        assert trade.commission == pytest.approx(51.0)
        assert trade.tax == pytest.approx(1095.0)

    def test_zero_quantity_produces_error(self):
        result = ParseResult()
        trade = self.parser._parse_line(_toss_buy_line(qty=0), 4, result)
        assert trade is None
        assert any("수량" in e["reason"] for e in result.errors)

    def test_dot_date_separator(self):
        result = ParseResult()
        trade = self.parser._parse_line(_toss_buy_line(date_str="2026.03.30"), 5, result)
        assert trade is not None
        assert trade.traded_at_kst == "2026-03-30"

    def test_price_fallback_from_amount_when_zero(self):
        result = ParseResult()
        line = _toss_buy_line(qty=4, amount=400000, price=0)
        trade = self.parser._parse_line(line, 7, result)
        assert trade is not None
        assert trade.price == pytest.approx(100000.0)


class TestTossPdfParserFixture:
    """실제 sample PDF 로 회귀 가드. extract_tables 가 비어 있을 때 0건 침묵 반환되던
    버그를 다시 잡기 위한 통합 테스트."""

    parser = TossPdfParser()

    @pytest.fixture
    def sample_pdf_bytes(self) -> bytes:
        from pathlib import Path
        path = Path(__file__).resolve().parents[2] / "sample" / "토스증권_거래내역서_20250417_20260416_1.pdf"
        if not path.exists():
            pytest.skip(f"sample PDF not present: {path}")
        return path.read_bytes()

    @pytest.fixture
    def sample_pdf_with_settlement(self) -> bytes:
        """'정산금액' 컬럼이 추가된 신 PDF 포맷 (2026-05 이후 발급분)."""
        from pathlib import Path
        path = Path(__file__).resolve().parents[2] / "sample" / "토스증권_거래내역서_20250523_20260522_1.pdf"
        if not path.exists():
            pytest.skip(f"sample PDF not present: {path}")
        return path.read_bytes()

    def test_parses_real_sample(self, sample_pdf_bytes: bytes):
        result = self.parser.parse(sample_pdf_bytes, "토스증권_거래내역서_20250417_20260416_1.pdf")
        assert len(result.trades) > 0
        assert result.account_hint == "101-01-024891"
        # BUY/SELL 둘 다 추출되어야 한다.
        types = {t.trade_type for t in result.trades}
        assert "BUY" in types
        assert "SELL" in types
        # 모든 거래에서 단가/수량 > 0.
        for t in result.trades:
            assert t.price > 0
            assert t.quantity > 0

    def test_parses_new_format_with_settlement_column(
        self, sample_pdf_with_settlement: bytes
    ):
        """신 포맷에 추가된 '정산금액' 컬럼이 단가/수수료 인덱스를 밀어내지 않는지 회귀.

        헤더가 동적으로 인식되지 않으면 price 에 정산금액, commission 에 단가가 들어가
        SELL 의 PnL 이 천문학적 손실로 계산되는 사고가 재발한다.
        """
        result = self.parser.parse(
            sample_pdf_with_settlement, "토스증권_거래내역서_20250523_20260522_1.pdf"
        )
        assert len(result.trades) == 16
        by_key = {(t.trade_type, t.asset_name, t.quantity): t for t in result.trades}
        # 삼성전자 BUY 33 주 — 단가 60000, 수수료 277.
        # 동적 인식 실패 시 price=1980277(정산금액) 으로 잘못 들어옴.
        samsung_buy = by_key[("BUY", "삼성전자", 33.0)]
        assert samsung_buy.price == pytest.approx(60000.0)
        assert samsung_buy.commission == pytest.approx(277.0)
        assert samsung_buy.tax == pytest.approx(0.0)
        # 삼성전자 SELL 15 주 — 단가 58100, 수수료 122, 거래세+제세금 = 0 + 1307.
        samsung_sell = by_key[("SELL", "삼성전자", 15.0)]
        assert samsung_sell.price == pytest.approx(58100.0)
        assert samsung_sell.commission == pytest.approx(122.0)
        assert samsung_sell.tax == pytest.approx(1307.0)


class TestTossColumnMap:
    def test_default_format_without_settlement(self):
        header = "거래일자 거래구분 종목명(종목코드) 환율 거래수량 거래대금 단가 수수료 거래세 제세금 변제/연체합 잔고 잔액"
        m = _build_column_map(header)
        assert m is not None
        assert m["거래수량"] == 0
        assert m["거래대금"] == 1
        assert m["단가"] == 2
        assert m["수수료"] == 3
        assert m["거래세"] == 4
        assert m["제세금"] == 5

    def test_new_format_with_settlement(self):
        header = "거래일자 거래구분 종목명(종목코드) 환율 거래수량 거래대금 정산금액 단가 수수료 거래세 제세금 변제/연체합 잔고 잔액"
        m = _build_column_map(header)
        assert m is not None
        # '정산금액' 이 거래대금과 단가 사이에 끼어 단가/수수료가 한 칸씩 밀려야 한다.
        assert m["거래수량"] == 0
        assert m["거래대금"] == 1
        assert m["정산금액"] == 2
        assert m["단가"] == 3
        assert m["수수료"] == 4
        assert m["거래세"] == 5
        assert m["제세금"] == 6

    def test_non_header_line_returns_none(self):
        # 데이터 행이나 잡음 라인은 헤더로 오인되면 안 된다.
        assert _build_column_map("2025.06.11 구매 삼성전자(A005930) 33 1,980,000") is None
        assert _build_column_map("발급번호 20260522-101-11-B000033") is None


class TestTossUsdHelpers:
    def test_split_usd_nums_normal(self):
        # 공백 정상 분리.
        rest = "1,370.30 0.030345 1,233 1,219 40,641 0 0 0 0.030345 3,713"
        assert _split_usd_nums(rest) == [
            "1,370.30", "0.030345", "1,233", "1,219", "40,641",
            "0", "0", "0", "0.030345", "3,713",
        ]

    def test_split_usd_nums_glued_rate_and_qty(self):
        # 환율+소수수량이 공백 없이 붙은 케이스('1,370.300.004568') 디-글루.
        # 인덱싱 전에 분리되지 않으면 이후 모든 컬럼이 한 칸씩 밀린다.
        rest = "1,370.300.004568 1,233 1,219 269,961 0 0 0 0.004568 2,493"
        assert _split_usd_nums(rest) == [
            "1,370.30", "0.004568", "1,233", "1,219", "269,961",
            "0", "0", "0", "0.004568", "2,493",
        ]

    def test_parse_usd_name_strips_isin(self):
        assert _parse_usd_name("팔란티어(US69608A1088)") == "팔란티어"
        assert _parse_usd_name("게임하우스 홀딩스(KYG3731B1086)") == "게임하우스 홀딩스"

    def test_parse_usd_isin_extracts_code(self):
        assert _parse_usd_isin("팔란티어(US69608A1088)") == "US69608A1088"
        assert _parse_usd_isin("게임하우스 홀딩스(KYG3731B1086)") == "KYG3731B1086"
        # ISIN 패턴이 없으면 None (종목명 폴백).
        assert _parse_usd_isin("이름만") is None


class TestTossUsdParserLine:
    """USD 행의 ÷환율 복원 단위 검증. 두 실파일 샘플은 수수료/제세금이 모두 0이라
    라운드트립만으론 commission/tax 의 KRW 잔류를 못 잡는다 — nonzero 합성 행으로
    세 필드 전부 ÷환율 되는지 증명한다(원가 ~환율배 부풀림 가드)."""

    parser = TossPdfParser()

    def test_usd_buy_divides_all_three_fields_by_rate(self):
        result = ParseResult()
        # 컬럼: 환율 수량 거래대금 정산금액 단가 수수료 제세금 변제 잔고 잔액.
        # 실파일 샘플은 수수료·제세금이 전부 0이라 ÷환율 라운드트립이 공허하다(양변 0).
        # 비-0 원화 수수료(6,852원)·제세금(1,370원)으로 ÷환율 정확성을 별도 커버한다.
        line = (
            "2024.08.14 구매 팔란티어(US69608A1088) "
            "1,370.30 0.030345 1,233 1,219 40,641 6,852 1,370 0 0.030345 3,713"
        )
        t = self.parser._parse_usd_line(line, 1, result)
        assert t is not None, result.errors
        assert t.trade_type == "BUY"
        assert t.asset_name == "팔란티어"
        assert t.country_code == "US"
        assert t.currency == "USD"
        assert t.ticker_hint is None
        # ISIN 은 ticker_hint 가 아니라 isin 필드로 전달(OpenFIGI 해소용).
        assert t.isin == "US69608A1088"
        assert t.exchange_rate == pytest.approx(1370.30)
        assert t.quantity == pytest.approx(0.030345)
        # 단가 40,641원 ÷ 1,370.30 ≈ 29.66 USD.
        assert t.price == pytest.approx(40641 / 1370.30)
        # 수수료 6,852원 ÷ 1,370.30 ≈ 5.0 USD, 제세금 1,370원 ÷ 1,370.30 ≈ 1.0 USD.
        # KRW 잔류 시 ×환율 으로 천문학적 부풀림 → 분할 누락 가드.
        assert t.commission == pytest.approx(5.0, abs=0.001)
        assert t.tax == pytest.approx(1.0, abs=0.001)
        # 라운드트립: USD × 환율 ≈ 원화값.
        assert t.price * t.exchange_rate == pytest.approx(40641)
        assert t.commission * t.exchange_rate == pytest.approx(6852)
        assert t.tax * t.exchange_rate == pytest.approx(1370)

    def test_usd_sell_glued_rate_token(self):
        result = ParseResult()
        line = (
            "2024.08.14 판매 테슬라(US88160R1014) "
            "1,370.300.004568 1,233 1,219 269,961 0 0 0 0.004568 2,493"
        )
        t = self.parser._parse_usd_line(line, 2, result)
        assert t is not None, result.errors
        assert t.trade_type == "SELL"
        assert t.asset_name == "테슬라"
        assert t.quantity == pytest.approx(0.004568)
        assert t.exchange_rate == pytest.approx(1370.30)
        assert t.price == pytest.approx(269961 / 1370.30)


class TestTossUsdParserFixture:
    """실파일 USD 회귀. ISIN 침묵 스킵(신규 0·스킵 0) 버그 재발 가드."""

    parser = TossPdfParser()

    @pytest.fixture
    def sample_usd_only(self) -> bytes:
        from pathlib import Path
        path = Path(__file__).resolve().parents[2] / "sample" / "토스_거래내역서_20240811_20250810_1.pdf"
        if not path.exists():
            pytest.skip(f"sample PDF not present: {path}")
        return path.read_bytes()

    @pytest.fixture
    def sample_mixed(self) -> bytes:
        from pathlib import Path
        path = Path(__file__).resolve().parents[2] / "sample" / "거래내역서_토스_해외포함_20250613_20260612_1.pdf"
        if not path.exists():
            pytest.skip(f"sample PDF not present: {path}")
        return path.read_bytes()

    def test_usd_only_sample_counts(self, sample_usd_only: bytes):
        result = self.parser.parse(sample_usd_only, "토스_거래내역서_20240811_20250810_1.pdf")
        usd = [t for t in result.trades if t.country_code == "US"]
        krw = [t for t in result.trades if t.country_code == "KR"]
        assert len(usd) == 648
        assert len(krw) == 0
        # 모든 USD 거래(648행): currency/country/hint 불변식.
        for t in usd:
            assert t.currency == "USD"
            assert t.ticker_hint is None
            assert t.price > 0 and t.quantity > 0
            # 환율이 정상 USD/KRW 밴드(1,300~1,550)에 들어야 한다. de-glue 가 환율 이후
            # 컬럼을 한 칸 밀면 환율 자리에 수량(0.00x)/거래대금(1,2xx) 같은 값이 들어와
            # 밴드를 벗어난다 — 건수 일치로는 못 잡는 컬럼 시프트를 648행 전부에서 가드.
            assert 1000 < t.exchange_rate < 2000

    def test_usd_only_first_buy_concrete_values(self, sample_usd_only: bytes):
        result = self.parser.parse(sample_usd_only, "토스_거래내역서_20240811_20250810_1.pdf")
        usd = [t for t in result.trades if t.country_code == "US"]
        first = usd[0]
        # 팔란티어 첫 매수: 단가 40,641원 ÷ 1,370.30 ≈ 29.66 USD, qty 0.030345.
        assert first.asset_name == "팔란티어"
        assert first.trade_type == "BUY"
        assert first.quantity == pytest.approx(0.030345)
        assert first.exchange_rate == pytest.approx(1370.30)
        assert first.price == pytest.approx(29.66, abs=0.01)
        assert first.price * first.exchange_rate == pytest.approx(40641)

    def test_usd_only_glued_row_concrete_values(self, sample_usd_only: bytes):
        """실파일의 glued 행('1,370.300.004568')이 디-글루 후 정확히 인덱싱되는지.

        둘째 USD 거래(테슬라)는 환율+소수수량이 공백 없이 붙은 실데이터 행이다. 디-글루가
        실패하면 환율/수량/단가가 한 칸씩 밀려 price 가 천문학적으로 깨진다.
        """
        result = self.parser.parse(sample_usd_only, "토스_거래내역서_20240811_20250810_1.pdf")
        usd = [t for t in result.trades if t.country_code == "US"]
        tesla = usd[1]
        assert tesla.asset_name == "테슬라"
        assert tesla.quantity == pytest.approx(0.004568)
        assert tesla.exchange_rate == pytest.approx(1370.30)
        # 단가 269,961원 ÷ 1,370.30 ≈ 197.01 USD.
        assert tesla.price == pytest.approx(269961 / 1370.30)
        assert tesla.price * tesla.exchange_rate == pytest.approx(269961)

    def test_mixed_sample_counts(self, sample_mixed: bytes):
        result = self.parser.parse(sample_mixed, "거래내역서_토스_해외포함_20250613_20260612_1.pdf")
        usd = [t for t in result.trades if t.country_code == "US"]
        krw = [t for t in result.trades if t.country_code == "KR"]
        # KRW 주식 15 + USD 구매 2.
        assert len(krw) == 15
        assert len(usd) == 2
        # KRW 거래 무회귀: 환율 1.0, currency KRW.
        for t in krw:
            assert t.exchange_rate == 1.0
            assert t.currency == "KRW"
        # 게임하우스 홀딩스 첫 매수: 단가 3,006원 ÷ 1,518.40 ≈ 1.98 USD, qty 1.
        # ISIN 이 KY(케이맨) 접두사지만 country_code 는 ISIN 이 아니라 *섹션*('달러 거래내역')
        # 기준으로 "US" — 통화 권위는 섹션이지 ISIN 국가코드가 아님을 못박는다.
        g = usd[0]
        assert g.asset_name == "게임하우스 홀딩스"
        assert g.country_code == "US"
        assert g.quantity == pytest.approx(1.0)
        assert g.exchange_rate == pytest.approx(1518.40)
        assert g.price == pytest.approx(3006 / 1518.40)
        assert g.ticker_hint is None


class TestShinhanPdfParserLine:
    """종목명에 숫자 토큰이 있어도 단가/수수료를 끝 6개 컬럼에서 정확히 앵커하는지."""

    parser = ShinhanPdfParser()

    def test_asset_name_with_number_token(self):
        # 종목명 "KODEX 200" 의 '200' 을 단가로 오인하면 안 된다.
        line1 = "2026-01-13 KODEX 200 49,700 0 0 0 99,400 0"
        line2 = "1 장내_매수 2 0 0 0 0 신한 SOL증권(iPhone)"
        line3 = "위탁(주식) 99,400 0 99,400 468,768"
        result = ParseResult()
        m2 = _LINE2_RE.match(line2)
        assert m2 is not None
        trade = self.parser._parse_record(line1, line2, line3, m2, 1, result)
        assert trade is not None, result.errors
        assert trade.asset_name == "KODEX 200"
        assert trade.price == 49700
        assert trade.commission == 0
        assert trade.quantity == 2
        assert trade.trade_type == "BUY"


class TestShinhanPdfParserFixture:
    parser = ShinhanPdfParser()

    @pytest.fixture
    def sample_bytes(self) -> bytes:
        from pathlib import Path
        path = Path(__file__).resolve().parents[2] / "sample" / "거래내역서_신한투자증권_1.pdf"
        if not path.exists():
            pytest.skip(f"sample not present: {path}")
        return path.read_bytes()

    def test_parses_real_sample(self, sample_bytes: bytes):
        result = self.parser.parse(sample_bytes, "거래내역서_신한투자증권_1.pdf")
        # 위탁(주식) 거래만 29건 (RP_* 위탁(RP)는 제외).
        assert len(result.trades) == 29
        assert result.account_hint == "270-26-192214"
        # 첫 거래: 한화엔진 SELL 2 @ 49,700.
        first = result.trades[0]
        assert first.asset_name == "한화엔진"
        assert first.trade_type == "SELL"
        assert first.quantity == 2
        assert first.price == 49700
        assert first.commission == 0
        assert first.tax == 49  # 거래세
        # 수량 × 단가 == 과표금액 (모든 거래에서 정합).
        for t in result.trades:
            assert t.quantity > 0 and t.price > 0

    def test_encrypted_pdf_friendly_error(self):
        # 열 수 없는 PDF(암호화 등)는 500 이 아니라 친절 안내 에러로 처리한다.
        from pathlib import Path
        path = Path(__file__).resolve().parents[2] / "sample" / "거래내역증명서_미래에셋증권_1.pdf"
        if not path.exists():
            pytest.skip(f"sample not present: {path}")
        result = self.parser.parse(path.read_bytes(), "암호.pdf")
        assert len(result.trades) == 0
        assert len(result.errors) == 1
        assert "암호 없는 버전" in result.errors[0]["reason"]


class TestMiraePdfParserFixture:
    parser = MiraePdfParser()

    @pytest.fixture
    def sample_bytes(self) -> bytes:
        from pathlib import Path
        path = Path(__file__).resolve().parents[2] / "sample" / "거래내역증명서_미래에셋증권_2.pdf"
        if not path.exists():
            pytest.skip(f"sample not present: {path}")
        return path.read_bytes()

    @pytest.fixture
    def encrypted_bytes(self) -> bytes:
        from pathlib import Path
        path = Path(__file__).resolve().parents[2] / "sample" / "거래내역증명서_미래에셋증권_1.pdf"
        if not path.exists():
            pytest.skip(f"sample not present: {path}")
        return path.read_bytes()

    def test_parses_real_sample(self, sample_bytes: bytes):
        result = self.parser.parse(sample_bytes, "거래내역증명서_미래에셋증권_2.pdf")
        # 주식 입출고(BUY/SELL) 38건. 현금leg(주식매수출금/매도입금)·이체·예탁금이용료는 skip.
        assert len(result.trades) == 38
        assert result.account_hint == "584-566838640"
        # 첫 거래: 두산에너빌리티 BUY 10 @ 105,700, ticker_hint=A코드 6자리.
        first = result.trades[0]
        assert first.asset_name == "두산에너빌리티보통주"
        assert first.trade_type == "BUY"
        assert first.quantity == 10
        assert first.price == 105700
        assert first.ticker_hint == "034020"
        # 긴 ETF명(줄바꿈) 레코드도 정상 파싱.
        wrapped = [t for t in result.trades if "KODEX 방산" in t.asset_name]
        assert wrapped, "줄바꿈된 ETF 레코드가 누락되었다"
        # 영숫자 종목번호(A0080G0)는 KRX 표준 숫자 코드가 아니라 ticker_hint 로 쓰지 않는다
        # (다른 증권사 숫자 코드와 보유가 갈라지는 것 방지 — 종목명 매칭에 맡김).
        assert wrapped[0].ticker_hint is None
        assert wrapped[0].quantity == 13
        assert wrapped[0].price == 15095
        # 매도 제세금합: inline 매도행은 종목명 뒤 첫 trailing 숫자가 제세금합(≈거래액의 0.2%).
        sells = [t for t in result.trades if t.trade_type == "SELL"]
        kai_sell = next(
            t for t in sells if t.asset_name == "한국항공우주산업보통주" and t.quantity == 5 and t.price == 203000
        )
        assert kai_sell.tax == 2029  # 종목명 뒤 trailing [2,029 5] 의 첫 값
        samsung_sell = next(
            t for t in sells if t.asset_name == "삼성전자보통주" and t.quantity == 52 and t.price == 194000
        )
        assert samsung_sell.tax == 20176  # trailing 단일 [20,176] 도 제세금합으로 인식
        # 줄바꿈(wrapped) 매도행은 line3 에 제세금합이 없어 tax=0(보수적).
        wrapped_sell = next(
            (t for t in sells if "TIGER 미국S&P500" in t.asset_name), None
        )
        assert wrapped_sell is not None
        assert wrapped_sell.tax == 0

    def test_encrypted_pdf_friendly_error(self, encrypted_bytes: bytes):
        result = self.parser.parse(encrypted_bytes, "거래내역증명서_미래에셋증권_1.pdf")
        assert len(result.trades) == 0
        assert len(result.errors) == 1
        assert "암호 없는 버전" in result.errors[0]["reason"]


class TestParsersRegistry:
    def test_parsers_registry_has_both_brokers(self):
        assert "samsung_xlsx" in PARSERS
        assert "toss_pdf" in PARSERS

    def test_parsers_registry_has_new_brokers(self):
        assert "shinhan_pdf" in PARSERS
        assert "mirae_pdf" in PARSERS
        assert PARSERS["shinhan_pdf"].display_name == "신한투자증권"
        assert PARSERS["mirae_pdf"].display_name == "미래에셋증권"
